"""
Excel exports for Dux Voucher reports — Ledger, Day Book, Cash & Bank Book.

Each whitelisted endpoint reuses the matching data function in
`reports_api.py`, builds a styled openpyxl workbook, and streams it back
as a binary download via `frappe.local.response`.

The look-and-feel mirrors the on-screen reports: dark navy header band,
zebra rows, blue Opening / Closing balance rows, totals row with a heavy
top border, red Dr / green Cr columns, INR number formatting.
"""

import io

import frappe
from frappe import _
from frappe.utils import flt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Colour palette (matches the on-screen UI) ────────────────────────
NAVY        = "0F172A"
HEADER_BG   = "1E293B"
ZEBRA_BG    = "FAFBFC"
OB_BG       = "DBEAFE"
OB_BORDER   = "93C5FD"
TOT_BG      = "F1F5F9"
DR_RED      = "DC2626"
CR_GREEN    = "16A34A"
GRAY_TXT    = "64748B"
SOFT_GRAY   = "94A3B8"
PERIOD_BG   = "F8FAFC"
ACCENT_LEDGER = "2563EB"
ACCENT_CASH   = "059669"
ACCENT_DAY    = "7C3AED"
ACCENT_STUDENT = "0D9488"

THIN       = Side(style="thin",   color="E2E8F0")
HAIR       = Side(style="hair",   color="F1F5F9")
TOTAL_TOP  = Side(style="medium", color=NAVY)

NUM_FMT     = '#,##0.00'
NUM_FMT_DR  = '#,##0.00" Dr"'
NUM_FMT_CR  = '#,##0.00" Cr"'


# ══════════════════════════════════════════════════════════════════════
# WHITELISTED ENDPOINTS
# ══════════════════════════════════════════════════════════════════════

@frappe.whitelist()
def export_ledger_xlsx(company, account, from_date, to_date,
                       party=None, party_type=None):
    from dux_voucher.dux_voucher.api.reports_api import get_ledger_statement
    d = get_ledger_statement(company, account, from_date, to_date,
                              party=party, party_type=party_type)
    wb = _build_ledger_workbook(d, "Ledger Statement",
                                 debit_label="Debit",
                                 credit_label="Credit",
                                 accent_color=ACCENT_LEDGER)
    label = d.get("account_name") or account
    fname = f"Ledger_{_safe(label)}_{from_date}_to_{to_date}.xlsx"
    _stream(wb, fname)


@frappe.whitelist()
def export_cash_bank_book_xlsx(company, account, from_date, to_date):
    from dux_voucher.dux_voucher.api.reports_api import get_ledger_statement
    d = get_ledger_statement(company, account, from_date, to_date)
    is_bank = d.get("account_type") == "Bank"
    title = "Bank Book" if is_bank else "Cash Book"
    wb = _build_ledger_workbook(d, title,
                                 debit_label="Receipts",
                                 credit_label="Payments",
                                 accent_color=ACCENT_CASH)
    label = d.get("account_name") or account
    fname = f"{title.replace(' ', '')}_{_safe(label)}_{from_date}_to_{to_date}.xlsx"
    _stream(wb, fname)


@frappe.whitelist()
def export_student_ledger_xlsx(company, kind, student, from_date, to_date):
    from dux_voucher.dux_voucher.api.reports_api import get_student_ledger
    d = get_student_ledger(company, kind, student, from_date, to_date)
    wb = _build_ledger_workbook(d, "Student Ledger",
                                 debit_label="Debit",
                                 credit_label="Credit",
                                 accent_color=ACCENT_STUDENT)
    label = d.get("account_name") or student
    fname = f"StudentLedger_{_safe(label)}_{from_date}_to_{to_date}.xlsx"
    _stream(wb, fname)


@frappe.whitelist()
def export_day_book_xlsx(company, from_date, to_date, voucher_type_filter=None):
    from dux_voucher.dux_voucher.api.reports_api import get_day_book
    d = get_day_book(company, from_date, to_date,
                      voucher_type_filter=voucher_type_filter)
    wb = _build_day_book_workbook(d, voucher_type_filter or "All")
    fname = f"DayBook_{_safe(company)}_{from_date}_to_{to_date}.xlsx"
    _stream(wb, fname)


# ══════════════════════════════════════════════════════════════════════
# WORKBOOK BUILDERS
# ══════════════════════════════════════════════════════════════════════

def _build_ledger_workbook(d, title, debit_label, credit_label, accent_color):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    # (label, width, h-align, kind)
    cols = [
        ("Date",                  14, "left",  None),
        ("Particulars",           42, "left",  None),
        ("Vch Type",              18, "left",  None),
        ("Vch No",                22, "left",  None),
        (f"{debit_label} (₹)",    16, "right", "amt"),
        (f"{credit_label} (₹)",   16, "right", "amt"),
        ("Balance (₹)",           18, "right", "bal"),
    ]
    n_cols = len(cols)

    sub = f"{title}: {d.get('account_name') or d.get('account')}"
    if d.get("account_type"):
        sub += f"  [{d['account_type']}]"
    period = (f"{d['from_date']}   →   {d['to_date']}    ·    "
              f"{d['row_count']} transaction(s)")

    header_row = _write_title_block(ws, d['company'], sub, period,
                                      n_cols, accent_color)
    _write_header_row(ws, [c[0] for c in cols], header_row,
                       aligns=[c[2] for c in cols])
    _apply_widths(ws, [c[1] for c in cols])
    ws.freeze_panes = f"A{header_row + 1}"

    row = header_row + 1

    # Opening Balance — figure on its natural Dr/Cr side
    row = _write_balance_row(ws, row, "Opening Balance",
                              d['opening_balance'], d['opening_type'], n_cols,
                              col_debit=d.get('opening_debit'),
                              col_credit=d.get('opening_credit'))

    # Data rows
    for i, r in enumerate(d['rows']):
        prefix = r.get('prefix') or ""
        contra = r.get('contra') or ""
        particulars = (prefix + "  " + contra) if prefix else contra
        zebra = (i % 2 == 1)

        ws.cell(row=row, column=1, value=r['posting_date'])
        ws.cell(row=row, column=2, value=particulars)
        ws.cell(row=row, column=3, value=r['voucher_type'])
        ws.cell(row=row, column=4, value=r['voucher_no'])

        dr = flt(r['debit'])
        cr = flt(r['credit'])
        ws.cell(row=row, column=5, value=(dr if dr else None))
        ws.cell(row=row, column=6, value=(cr if cr else None))
        ws.cell(row=row, column=7, value=flt(r['balance']))

        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal=cols[col_idx - 1][2],
                                         vertical="top",
                                         wrap_text=(col_idx == 2))
            cell.border = Border(bottom=HAIR)
            if zebra:
                cell.fill = PatternFill("solid", fgColor=ZEBRA_BG)

        # Number formats + colours
        ws.cell(row=row, column=5).number_format = NUM_FMT
        ws.cell(row=row, column=6).number_format = NUM_FMT
        bal_cell = ws.cell(row=row, column=7)
        if r['balance_type'] == 'Dr':
            bal_cell.number_format = NUM_FMT_DR
            bal_cell.font = Font(name="Calibri", size=10, bold=True, color=DR_RED)
        else:
            bal_cell.number_format = NUM_FMT_CR
            bal_cell.font = Font(name="Calibri", size=10, bold=True, color=CR_GREEN)
        if dr:
            ws.cell(row=row, column=5).font = Font(name="Calibri", size=10, color=DR_RED)
        if cr:
            ws.cell(row=row, column=6).font = Font(name="Calibri", size=10, color=CR_GREEN)

        row += 1

        # Drill-down sub-rows for a "Various" particular — the individual
        # heads that make it up, indented under the row, each on its side.
        for b in (r.get('breakdown') or []):
            for col_idx in range(1, n_cols + 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.border = Border(bottom=HAIR)
                if zebra:
                    cell.fill = PatternFill("solid", fgColor=ZEBRA_BG)
            lc = ws.cell(row=row, column=2, value="↳  " + (b.get('label') or ""))
            lc.font = Font(name="Calibri", size=9, italic=True, color=GRAY_TXT)
            lc.alignment = Alignment(horizontal="left", vertical="top",
                                       wrap_text=True, indent=2)
            is_cr = (b.get('side') == 'Cr')
            ac = ws.cell(row=row, column=(6 if is_cr else 5), value=flt(b.get('amount')))
            ac.font = Font(name="Calibri", size=9,
                            color=(CR_GREEN if is_cr else DR_RED))
            ac.alignment = Alignment(horizontal="right", vertical="top")
            ac.number_format = NUM_FMT
            row += 1

        # Remarks (italic, gray, merged across particulars→last col)
        if r.get('remarks'):
            cell = ws.cell(row=row, column=2, value=r['remarks'])
            cell.font = Font(name="Calibri", size=9, italic=True, color=SOFT_GRAY)
            cell.alignment = Alignment(horizontal="left", vertical="top",
                                         wrap_text=True, indent=1)
            ws.merge_cells(start_row=row, start_column=2,
                            end_row=row, end_column=n_cols)
            if zebra:
                fill = PatternFill("solid", fgColor=ZEBRA_BG)
                for c in range(1, n_cols + 1):
                    ws.cell(row=row, column=c).fill = fill
            row += 1

    # Closing Balance — carried down as a balancing contra (opposite side)
    row = _write_balance_row(ws, row, "Closing Balance",
                              d['closing_balance'], d['closing_type'], n_cols,
                              col_debit=d.get('closing_balancing_debit'),
                              col_credit=d.get('closing_balancing_credit'))

    # Total — balanced grand total (Debit == Credit)
    row = _write_totals_row(ws, row, n_cols,
                              total_debit=d['grand_total_debit'],
                              total_credit=d['grand_total_credit'])

    # Footer
    row += 1
    foot = ws.cell(row=row, column=n_cols, value="Generated by Dux DigiTech")
    foot.font = Font(name="Calibri", size=9, italic=True, color=SOFT_GRAY)
    foot.alignment = Alignment(horizontal="right")

    _apply_page_setup(ws, header_row, landscape=True)
    return wb


def _build_day_book_workbook(d, vt_filter):
    wb = Workbook()
    ws = wb.active
    ws.title = "Day Book"

    cols = [
        ("Date",         14, "left",  None),
        ("Vch Type",     18, "left",  None),
        ("Vch No",       22, "left",  None),
        ("Particulars",  50, "left",  None),
        ("Amount (₹)",   18, "right", "amt"),
    ]
    n_cols = len(cols)

    sub = "Day Book"
    if vt_filter and vt_filter != "All":
        sub += f"  ·  Filter: {vt_filter}"
    period = (f"{d['from_date']}   →   {d['to_date']}    ·    "
              f"{d['row_count']} voucher(s)")

    header_row = _write_title_block(ws, d['company'], sub, period,
                                      n_cols, ACCENT_DAY)
    _write_header_row(ws, [c[0] for c in cols], header_row,
                       aligns=[c[2] for c in cols])
    _apply_widths(ws, [c[1] for c in cols])
    ws.freeze_panes = f"A{header_row + 1}"

    row = header_row + 1

    for i, r in enumerate(d['rows']):
        zebra = (i % 2 == 1)
        ws.cell(row=row, column=1, value=r['posting_date'])
        ws.cell(row=row, column=2, value=r['voucher_type'])
        ws.cell(row=row, column=3, value=r['voucher_no'])
        ws.cell(row=row, column=4, value=r['particulars'])
        ws.cell(row=row, column=5, value=flt(r['amount']))

        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal=cols[col_idx - 1][2],
                                         vertical="top",
                                         wrap_text=(col_idx == 4))
            cell.border = Border(bottom=HAIR)
            if zebra:
                cell.fill = PatternFill("solid", fgColor=ZEBRA_BG)

        ws.cell(row=row, column=5).number_format = NUM_FMT
        row += 1

        if r.get('remarks'):
            cell = ws.cell(row=row, column=4, value=r['remarks'])
            cell.font = Font(name="Calibri", size=9, italic=True, color=SOFT_GRAY)
            cell.alignment = Alignment(horizontal="left", vertical="top",
                                         wrap_text=True, indent=1)
            ws.merge_cells(start_row=row, start_column=4,
                            end_row=row, end_column=n_cols)
            if zebra:
                fill = PatternFill("solid", fgColor=ZEBRA_BG)
                for c in range(1, n_cols + 1):
                    ws.cell(row=row, column=c).fill = fill
            row += 1

    # Totals — one amount column for Day Book
    fill = PatternFill("solid", fgColor=TOT_BG)
    border = Border(top=TOTAL_TOP, bottom=THIN)
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.border = border

    label_cell = ws.cell(row=row, column=1,
                          value=f"TOTAL — {d['row_count']} voucher(s)")
    label_cell.font = Font(name="Calibri", size=10, bold=True, color=GRAY_TXT)
    label_cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(start_row=row, start_column=1,
                    end_row=row, end_column=n_cols - 1)

    amt_cell = ws.cell(row=row, column=n_cols, value=flt(d['total_debit']))
    amt_cell.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    amt_cell.alignment = Alignment(horizontal="right", vertical="center")
    amt_cell.number_format = NUM_FMT
    ws.row_dimensions[row].height = 22

    row += 2
    foot = ws.cell(row=row, column=n_cols, value="Generated by Dux DigiTech")
    foot.font = Font(name="Calibri", size=9, italic=True, color=SOFT_GRAY)
    foot.alignment = Alignment(horizontal="right")

    _apply_page_setup(ws, header_row, landscape=True)
    return wb


# ══════════════════════════════════════════════════════════════════════
# SHARED HELPERS
# ══════════════════════════════════════════════════════════════════════

def _write_title_block(ws, company, sub, period, n_cols, accent_color):
    """Three-row banner: company / report-type / period. Returns the
    row number where the table header should be written (row 5)."""
    c1 = ws.cell(row=1, column=1, value=company)
    c1.font = Font(name="Calibri", size=18, bold=True, color=NAVY)
    c1.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=n_cols)
    ws.row_dimensions[1].height = 28

    c2 = ws.cell(row=2, column=1, value=sub)
    c2.font = Font(name="Calibri", size=12, bold=True, color=accent_color)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=n_cols)
    ws.row_dimensions[2].height = 20

    c3 = ws.cell(row=3, column=1, value=period)
    c3.font = Font(name="Calibri", size=10, italic=True, color=GRAY_TXT)
    c3.alignment = Alignment(horizontal="center", vertical="center")
    c3.fill = PatternFill("solid", fgColor=PERIOD_BG)
    ws.merge_cells(start_row=3, end_row=3, start_column=1, end_column=n_cols)
    ws.row_dimensions[3].height = 18

    ws.row_dimensions[4].height = 6
    return 5


def _write_header_row(ws, labels, row_num, aligns=None):
    aligns = aligns or ["left"] * len(labels)
    for col_idx, label in enumerate(labels, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=label)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal=aligns[col_idx - 1],
                                     vertical="center")
        cell.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
    ws.row_dimensions[row_num].height = 24


def _apply_widths(ws, widths):
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def _write_balance_row(ws, row, label, balance, balance_type, n_cols,
                        col_debit=None, col_credit=None):
    """Opening / Closing balance row.

    The label spans the text columns (Date … Vch No); the Debit and Credit
    columns carry the figure on the appropriate side — opening on its
    NATURAL side, closing carried down as a BALANCING contra on the
    OPPOSITE side — and the rightmost Balance column shows the running
    balance with its Dr/Cr suffix.
    """
    fill = PatternFill("solid", fgColor=OB_BG)
    border = Border(top=Side(style="thin", color=OB_BORDER),
                     bottom=Side(style="thin", color=OB_BORDER))
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.border = border

    label_cell = ws.cell(row=row, column=1, value=label.upper())
    label_cell.font = Font(name="Calibri", size=10, bold=True, color=NAVY)
    label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    # Label spans the four text columns; Debit / Credit / Balance carry figures.
    ws.merge_cells(start_row=row, start_column=1,
                    end_row=row, end_column=n_cols - 3)

    if col_debit:
        dc = ws.cell(row=row, column=n_cols - 2, value=flt(col_debit))
        dc.font = Font(name="Calibri", size=10, bold=True, color=DR_RED)
        dc.alignment = Alignment(horizontal="right", vertical="center")
        dc.number_format = NUM_FMT
    if col_credit:
        cc = ws.cell(row=row, column=n_cols - 1, value=flt(col_credit))
        cc.font = Font(name="Calibri", size=10, bold=True, color=CR_GREEN)
        cc.alignment = Alignment(horizontal="right", vertical="center")
        cc.number_format = NUM_FMT

    bal_cell = ws.cell(row=row, column=n_cols, value=flt(balance))
    bal_cell.font = Font(name="Calibri", size=11, bold=True,
                          color=DR_RED if balance_type == "Dr" else CR_GREEN)
    bal_cell.alignment = Alignment(horizontal="right", vertical="center")
    bal_cell.number_format = NUM_FMT_DR if balance_type == "Dr" else NUM_FMT_CR

    ws.row_dimensions[row].height = 22
    return row + 1


def _write_totals_row(ws, row, n_cols, total_debit, total_credit):
    fill = PatternFill("solid", fgColor=TOT_BG)
    border = Border(top=TOTAL_TOP, bottom=THIN)
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.border = border

    label_cell = ws.cell(row=row, column=1, value="TOTAL")
    label_cell.font = Font(name="Calibri", size=10, bold=True, color=GRAY_TXT)
    label_cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(start_row=row, start_column=1,
                    end_row=row, end_column=n_cols - 3)

    dr = ws.cell(row=row, column=n_cols - 2, value=flt(total_debit))
    dr.font = Font(name="Calibri", size=10, bold=True, color=DR_RED)
    dr.alignment = Alignment(horizontal="right", vertical="center")
    dr.number_format = NUM_FMT

    cr = ws.cell(row=row, column=n_cols - 1, value=flt(total_credit))
    cr.font = Font(name="Calibri", size=10, bold=True, color=CR_GREEN)
    cr.alignment = Alignment(horizontal="right", vertical="center")
    cr.number_format = NUM_FMT

    ws.row_dimensions[row].height = 22
    return row + 1


def _apply_page_setup(ws, header_row, landscape=True):
    ws.page_setup.orientation = (ws.ORIENTATION_LANDSCAPE if landscape
                                  else ws.ORIENTATION_PORTRAIT)
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5


def _safe(s):
    s = str(s or "")
    for ch in r'\/:*?"<>|':
        s = s.replace(ch, "_")
    return s.replace(" ", "_")[:50]


def _stream(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    frappe.local.response.filename = filename
    frappe.local.response.filecontent = buf.getvalue()
    frappe.local.response.type = "binary"
