"""
TB v3 — Option C: split into two sheets.
  Sheet 1 'Summary'  — management dashboard view, KPI-style
  Sheet 2 'Detail'   — full 6-column TB (v2 minimalist layout)
Both sheets reference the same numbers via formulas, so they cannot drift.
Hyperlinks connect them in both directions.
"""
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SOURCE_FILE = '/home/claude/work/source.xlsx'
OUTPUT_FILE = '/home/claude/work/TB_JEWIPL_v5.xlsx'

COMPANY_NAME  = "JAIN ENGINEERING WORKS (INDIA) PRIVATE LIMITED"
COMPANY_SUFFIX = " - JEWIPL"
PERIOD_LABEL  = "Trial Balance  ·  FY 2026-2027  (1 Apr 2026 – 31 Mar 2027)"

INR_FMT  = '[>=10000000]##\\,##\\,##\\,##0.00;[>=100000]##\\,##\\,##0.00;##,##0.00'
INR_BIG  = '[>=10000000]##\\,##\\,##\\,##0;[>=100000]##\\,##\\,##0;##,##0'

SLATE      = '2C3E50'
SLATE_LITE = '7F8C9A'
SLATE_FAINT= 'EAEEF1'
LINK_BLUE  = '0563C1'
MUTED_RED  = 'B91C1C'
MUTED_GRN  = '047857'

# v4 additions — hierarchy fills + hairline rule colour
TOP_FILL   = 'E1E7EF'   # slate-tint (low saturation) for top-group rows
SUB_FILL   = 'F2F3F5'   # soft gray for sub-group rows
HAIR_GRAY  = 'D6DBE0'   # row separator hairline

# v5 additions — status banner palette
ALERT_BG     = 'FCEAEA'   # light red wash for out-of-balance banner
ALERT_BORDER = 'F0B5B5'   # red border for the banner
OK_BG        = 'E8F5EE'   # light green wash for tied banner (production)
OK_BORDER    = 'A7D5B8'   # green border (production)

# ============================================================================
# READ + PARSE SOURCE
# ============================================================================
src = load_workbook(SOURCE_FILE)
sws = src.active
raw = [tuple(r) for r in sws.iter_rows(values_only=True)]

def indent_level(s):
    if s is None: return 0
    return len(re.match(r'^( *)', s).group(1)) // 4

def clean_name(s):
    if s is None: return ''
    name = s.strip()
    if name.endswith(COMPANY_SUFFIX):
        name = name[: -len(COMPANY_SUFFIX)]
    return name

data = []
for r in raw[1:-1]:
    if r[0] is None: continue
    data.append({
        'level': indent_level(r[0]),
        'name': clean_name(r[0]),
        'nums': list(r[1:7]),
    })
for i, row in enumerate(data):
    nxt = data[i + 1] if i + 1 < len(data) else None
    row['is_group'] = bool(nxt and nxt['level'] > row['level'])

total_nums = list(raw[-1][1:7])

wb = Workbook()
wb.remove(wb.active)

# Borders
hair    = Side(style='hair',   color=SLATE_LITE)
thin    = Side(style='thin',   color=SLATE_LITE)
medium  = Side(style='medium', color=SLATE)
double  = Side(style='double', color=SLATE)
row_sep = Side(style='hair',   color=HAIR_GRAY)   # v4: subtle row separator

# ============================================================================
# SHEET 2 — DETAIL  (build first so Summary can reference exact row numbers)
# ============================================================================
ws_d = wb.create_sheet("Detail")
ws_d.sheet_properties.outlinePr.summaryBelow = False
ws_d.sheet_properties.outlinePr.summaryRight = False
ws_d.sheet_view.showGridLines = False

# Title block
ws_d['A1'] = COMPANY_NAME
ws_d['A1'].font = Font(name='Calibri', size=12, bold=True, color=SLATE)
ws_d['A1'].alignment = Alignment(horizontal='left', vertical='center')
ws_d.row_dimensions[1].height = 20

ws_d['A2'] = PERIOD_LABEL
ws_d['A2'].font = Font(name='Calibri', size=10, color=SLATE_LITE)
ws_d['A2'].alignment = Alignment(horizontal='left', vertical='center')
ws_d.row_dimensions[2].height = 15

# Back-to-Summary link (row 2, right side)
back = ws_d.cell(row=2, column=7, value='←  Back to Summary')
back.hyperlink = "#'Summary'!A1"
back.font = Font(name='Calibri', size=10, color=LINK_BLUE, underline='single')
back.alignment = Alignment(horizontal='right')

for col in range(1, 8):
    ws_d.cell(row=3, column=col).border = Border(bottom=medium)
ws_d.row_dimensions[3].height = 4

HDR = 4
for col, h in enumerate(
    ['Account', 'Opening (Dr)', 'Opening (Cr)',
     'Debit', 'Credit', 'Closing (Dr)', 'Closing (Cr)'], 1):
    c = ws_d.cell(row=HDR, column=col, value=h)
    c.font = Font(name='Calibri', size=10, bold=True, color=SLATE)
    c.alignment = Alignment(
        horizontal=('left' if col == 1 else 'right'),
        vertical='center')
    c.border = Border(bottom=thin)
ws_d.row_dimensions[HDR].height = 22

# Track each top-group's row number so Summary can hyperlink there
top_group_rows = {}

r = HDR + 1
for i, row in enumerate(data):
    is_top = (row['level'] == 0 and row['is_group'])
    if is_top:
        top_group_rows[row['name']] = r

    a = ws_d.cell(row=r, column=1, value=row['name'])
    a.alignment = Alignment(horizontal='left', vertical='center', indent=row['level'])

    for j, v in enumerate(row['nums'], 2):
        c = ws_d.cell(row=r, column=j, value=(None if v in (0, None) else v))
        c.number_format = INR_FMT
        c.alignment = Alignment(horizontal='right', vertical='center')

    if is_top:
        # Top group: slate-tint fill, slate text, bold
        fill = PatternFill('solid', fgColor=TOP_FILL)
        font = Font(name='Calibri', size=11, bold=True, color=SLATE)
        for col in range(1, 8):
            cc = ws_d.cell(row=r, column=col)
            cc.font = font
            cc.fill = fill
            cc.border = Border(bottom=row_sep)
        ws_d.row_dimensions[r].height = 22

    elif row['is_group'] and row['level'] == 1:
        # Sub-group (Current Assets, Current Liabilities, Direct Expenses…):
        # soft gray fill, bold black
        fill = PatternFill('solid', fgColor=SUB_FILL)
        for col in range(1, 8):
            cc = ws_d.cell(row=r, column=col)
            cc.font = Font(name='Calibri', size=10, bold=True)
            cc.fill = fill
            cc.border = Border(bottom=row_sep)
        ws_d.row_dimensions[r].height = 19

    elif row['is_group']:
        # Deeper groups (Accounts Receivable, Bank Accounts…): bold only
        for col in range(1, 8):
            cc = ws_d.cell(row=r, column=col)
            cc.font = Font(name='Calibri', size=10, bold=True)
            cc.border = Border(bottom=row_sep)

    else:
        # Leaf row
        for col in range(1, 8):
            cc = ws_d.cell(row=r, column=col)
            cc.font = Font(name='Calibri', size=10, color='3D3D3D')
            cc.border = Border(bottom=row_sep)

    ws_d.row_dimensions[r].outline_level = 0 if row['level'] == 0 else 1
    r += 1

total_r = r
ws_d.cell(row=total_r, column=1, value='Total')
for j, v in enumerate(total_nums, 2):
    c = ws_d.cell(row=total_r, column=j, value=(None if v in (0, None) else v))
    c.number_format = INR_FMT
for col in range(1, 8):
    cc = ws_d.cell(row=total_r, column=col)
    cc.font = Font(name='Calibri', size=11, bold=True, color=SLATE)
    cc.border = Border(top=medium, bottom=double)
    cc.alignment = Alignment(
        horizontal=('left' if col == 1 else 'right'), vertical='center')
ws_d.row_dimensions[total_r].height = 22

check_r = total_r + 2
ws_d.cell(row=check_r, column=1, value='Check  ·  Closing (Dr) − Closing (Cr)')
ws_d.cell(row=check_r, column=1).font = Font(
    name='Calibri', size=9, italic=True, color=SLATE_LITE)
diff_d = ws_d.cell(row=check_r, column=7,
                   value=f'=F{total_r}-G{total_r}')
diff_d.number_format = INR_FMT
diff_d.font = Font(name='Calibri', size=9, italic=True, color=MUTED_RED, bold=True)
diff_d.alignment = Alignment(horizontal='right')

ws_d.column_dimensions['A'].width = 42
for col in 'BCDEFG':
    ws_d.column_dimensions[col].width = 15
ws_d.freeze_panes = 'B5'
ws_d.print_title_rows = '1:4'
ws_d.page_setup.orientation = ws_d.ORIENTATION_LANDSCAPE
ws_d.page_setup.paperSize = ws_d.PAPERSIZE_A4
ws_d.page_setup.fitToWidth = 1
ws_d.page_setup.fitToHeight = 0
ws_d.sheet_properties.pageSetUpPr.fitToPage = True
ws_d.page_margins.left = 0.3; ws_d.page_margins.right = 0.3
ws_d.page_margins.top = 0.5;  ws_d.page_margins.bottom = 0.5

# ============================================================================
# SHEET 1 — SUMMARY  (placed first in tab order)
# ============================================================================
ws_s = wb.create_sheet("Summary", 0)
ws_s.sheet_view.showGridLines = False

# Pre-compute whether the TB ties — drives the banner styling at gen-time.
# (Production custom app will recompute this from the live data.)
tb_diff = (raw[-1][5] or 0) - (raw[-1][6] or 0)   # Closing Dr − Closing Cr
tb_tied = abs(tb_diff) < 0.01

# ---- Column widths (set early so merges respect them) ---------------------
ws_s.column_dimensions['A'].width = 3
ws_s.column_dimensions['B'].width = 34
ws_s.column_dimensions['C'].width = 2
ws_s.column_dimensions['D'].width = 18
ws_s.column_dimensions['E'].width = 18
ws_s.column_dimensions['F'].width = 3

# ---- TITLE BLOCK (rows 2–4) ------------------------------------------------
ws_s.merge_cells('B2:E2')
ws_s['B2'] = COMPANY_NAME
ws_s['B2'].font = Font(name='Calibri', size=14, bold=True, color=SLATE)
ws_s['B2'].alignment = Alignment(horizontal='left', vertical='center')
ws_s.row_dimensions[2].height = 26

ws_s.merge_cells('B3:E3')
ws_s['B3'] = PERIOD_LABEL
ws_s['B3'].font = Font(name='Calibri', size=10, color=SLATE_LITE)
ws_s['B3'].alignment = Alignment(horizontal='left', vertical='center')
ws_s.row_dimensions[3].height = 16

# Accent rule under title (B:E)
for col in range(2, 6):
    ws_s.cell(row=4, column=col).border = Border(bottom=medium)
ws_s.row_dimensions[4].height = 4

# ---- TIE STATUS BANNER  (rows 6-8) ----------------------------------------
# Card with fill + border. Red wash if out-of-balance, green if tied.
banner_bg     = OK_BG     if tb_tied else ALERT_BG
banner_border = OK_BORDER if tb_tied else ALERT_BORDER
banner_text   = MUTED_GRN if tb_tied else MUTED_RED
banner_icon   = '✓' if tb_tied else '⚠'
banner_head   = ('TRIAL BALANCE TIED' if tb_tied
                 else 'TRIAL BALANCE OUT OF BALANCE')
banner_sub    = ('Closing (Dr) equals Closing (Cr)' if tb_tied
                 else 'Difference  ·  Closing (Dr) − Closing (Cr)')

bn_side = Side(style='thin', color=banner_border)
bn_fill = PatternFill('solid', fgColor=banner_bg)

# Row 6: icon + headline, merged across B:E
ws_s.merge_cells('B6:E6')
banner = ws_s.cell(row=6, column=2, value=f'  {banner_icon}   {banner_head}')
banner.font = Font(name='Calibri', size=13, bold=True, color=banner_text)
banner.alignment = Alignment(horizontal='left', vertical='center')

# Row 7: sub-text + amount (only when out of balance shows the figure prominently)
ws_s.merge_cells('B7:D7')
sub = ws_s.cell(row=7, column=2, value=f'    {banner_sub}')
sub.font = Font(name='Calibri', size=10, color=SLATE)
sub.alignment = Alignment(horizontal='left', vertical='center')

amt = ws_s.cell(row=7, column=5,
                value=(0 if tb_tied
                       else f'=Detail!F{total_r}-Detail!G{total_r}'))
amt.number_format = INR_FMT
amt.font = Font(name='Calibri', size=12, bold=True, color=banner_text)
amt.alignment = Alignment(horizontal='right', vertical='center', indent=1)

# Apply fill + border across the whole banner area (B6:E7)
for r_ in (6, 7):
    for c_ in range(2, 6):
        cc = ws_s.cell(row=r_, column=c_)
        cc.fill = bn_fill
        # build perimeter border per cell
        top_b    = bn_side if r_ == 6 else None
        bot_b    = bn_side if r_ == 7 else None
        left_b   = bn_side if c_ == 2 else None
        right_b  = bn_side if c_ == 5 else None
        cc.border = Border(top=top_b, bottom=bot_b, left=left_b, right=right_b)

ws_s.row_dimensions[6].height = 26
ws_s.row_dimensions[7].height = 22

# ---- CLOSING POSITION (section starts row 10) -----------------------------
sr = 10
ws_s.cell(row=sr, column=2, value='Closing Position').font = Font(
    name='Calibri', size=11, bold=True, color=SLATE)
ws_s.row_dimensions[sr].height = 22
sr += 1

# Table header — slate-tint fill (matches Detail's top-group treatment)
hdr_fill = PatternFill('solid', fgColor=TOP_FILL)
ws_s.cell(row=sr, column=2, value='Group')
ws_s.cell(row=sr, column=4, value='Closing (Dr)')
ws_s.cell(row=sr, column=5, value='Closing (Cr)')
for c_ in range(2, 6):
    cc = ws_s.cell(row=sr, column=c_)
    cc.font = Font(name='Calibri', size=10, bold=True, color=SLATE)
    cc.fill = hdr_fill
    cc.border = Border(bottom=row_sep)
    if c_ in (4, 5):
        cc.alignment = Alignment(horizontal='right', vertical='center')
    else:
        cc.alignment = Alignment(horizontal='left', vertical='center')
ws_s.row_dimensions[sr].height = 20
sr += 1

# rows for each top group
top_groups_for_summary = [
    ('Application of Funds (Assets)',  'Assets'),
    ('Source of Funds (Liabilities)',  'Liabilities'),
    ('Income',                         'Income'),
    ('Expenses',                       'Expenses'),
]
for tg_name, display in top_groups_for_summary:
    if tg_name in top_group_rows:
        det_r = top_group_rows[tg_name]
        nm = ws_s.cell(row=sr, column=2, value=display)
        nm.hyperlink = f"#'Detail'!A{det_r}"
        nm.font = Font(name='Calibri', size=11, color=LINK_BLUE,
                       underline='single')
        cd = ws_s.cell(row=sr, column=4,
                       value=f'=IF(Detail!F{det_r}=0,"—",Detail!F{det_r})')
        cd.number_format = INR_FMT
        cd.font = Font(name='Calibri', size=11, color=SLATE)
        cd.alignment = Alignment(horizontal='right', vertical='center')
        cc = ws_s.cell(row=sr, column=5,
                       value=f'=IF(Detail!G{det_r}=0,"—",Detail!G{det_r})')
        cc.number_format = INR_FMT
        cc.font = Font(name='Calibri', size=11, color=SLATE)
        cc.alignment = Alignment(horizontal='right', vertical='center')
    else:
        nm = ws_s.cell(row=sr, column=2, value=display)
        nm.font = Font(name='Calibri', size=11, color=SLATE_LITE,
                       italic=True)
        for ccol, val in [(4, '—'), (5, '—')]:
            nm_d = ws_s.cell(row=sr, column=ccol, value=val)
            nm_d.font = Font(name='Calibri', size=11, color=SLATE_LITE)
            nm_d.alignment = Alignment(horizontal='right', vertical='center')
    # Hairline between rows
    for c_ in range(2, 6):
        ws_s.cell(row=sr, column=c_).border = Border(bottom=row_sep)
    ws_s.row_dimensions[sr].height = 22
    sr += 1

# ---- PERIOD ACTIVITY ------------------------------------------------------
sr += 2
ws_s.cell(row=sr, column=2, value='Period Activity').font = Font(
    name='Calibri', size=11, bold=True, color=SLATE)
ws_s.row_dimensions[sr].height = 22
sr += 1

# Table header
ws_s.cell(row=sr, column=2, value='Movement')
ws_s.cell(row=sr, column=4, value='Amount')
ws_s.merge_cells(start_row=sr, start_column=4, end_row=sr, end_column=5)
for c_ in (2, 4, 5):
    cc = ws_s.cell(row=sr, column=c_)
    cc.font = Font(name='Calibri', size=10, bold=True, color=SLATE)
    cc.fill = hdr_fill
    cc.border = Border(bottom=row_sep)
    if c_ == 2:
        cc.alignment = Alignment(horizontal='left', vertical='center')
    else:
        cc.alignment = Alignment(horizontal='right', vertical='center')
ws_s.row_dimensions[sr].height = 20
sr += 1

activity_rows = [
    ('Total Debits',       f'=Detail!D{total_r}', False),
    ('Total Credits',      f'=Detail!E{total_r}', False),
    ('Net (Dr − Cr)',      f'=Detail!D{total_r}-Detail!E{total_r}', True),
]
for lbl, formula, is_net in activity_rows:
    nm = ws_s.cell(row=sr, column=2, value=lbl)
    nm.font = Font(name='Calibri', size=11,
                   bold=is_net,
                   color=(banner_text if is_net else '000000'))
    nm.alignment = Alignment(horizontal='left', vertical='center')

    val = ws_s.cell(row=sr, column=4, value=formula)
    val.number_format = INR_FMT
    val.font = Font(name='Calibri', size=11,
                    bold=is_net,
                    color=(banner_text if is_net else SLATE))
    val.alignment = Alignment(horizontal='right', vertical='center')
    ws_s.merge_cells(start_row=sr, start_column=4, end_row=sr, end_column=5)

    for c_ in (2, 4, 5):
        cc_ = ws_s.cell(row=sr, column=c_)
        if is_net:
            cc_.border = Border(top=thin, bottom=row_sep)
            cc_.fill = PatternFill('solid', fgColor=banner_bg)
        else:
            cc_.border = Border(bottom=row_sep)
    ws_s.row_dimensions[sr].height = 22
    sr += 1

# ---- footer link to detail ------------------------------------------------
sr += 3
view_d = ws_s.cell(row=sr, column=2,
                   value='View full account-level detail  →')
view_d.hyperlink = "#'Detail'!A1"
view_d.font = Font(name='Calibri', size=10, color=LINK_BLUE,
                   underline='single')
view_d.alignment = Alignment(horizontal='left', vertical='center')

# Page setup
ws_s.page_setup.orientation = ws_s.ORIENTATION_PORTRAIT
ws_s.page_setup.paperSize = ws_s.PAPERSIZE_A4
ws_s.page_setup.fitToWidth = 1
ws_s.sheet_properties.pageSetUpPr.fitToPage = True
ws_s.page_margins.left = 0.4
ws_s.page_margins.right = 0.4
ws_s.page_margins.top = 0.5
ws_s.page_margins.bottom = 0.5

wb.active = 0  # Summary opens first

wb.save(OUTPUT_FILE)
print(f"Wrote: {OUTPUT_FILE}")
print(f"Detail total row: {total_r}")
print(f"Top groups found in detail: {list(top_group_rows.keys())}")
print(f"TB tied: {tb_tied}  Difference: {tb_diff}")
