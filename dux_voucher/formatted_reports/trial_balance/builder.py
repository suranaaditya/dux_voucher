"""
Build the v5 formatted Trial Balance workbook from ERPNext's
`trial_balance.execute()` output.

This module is purely about presentation. It receives `columns` and
`rows` exactly as ERPNext computed them and renders them into a
polished two-sheet xlsx. **No accounting logic lives here** — if the
numbers are wrong, the fix belongs in the standard report.

The visual contract is locked in `formatted_reports/PLAN.md` §3 (v5
spec). The reference implementation `build_v5_reference.py` (sibling
file) was the visual source-of-truth from the planning session.

Public entry point: :func:`build_formatted_tb`. Helpers prefixed with
``_`` are exposed for the unit-test module only.
"""

import io
import re
from datetime import date as _date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.hyperlink import Hyperlink


def _intra_link(cell, target_sheet, target_cell):
    """Attach an in-workbook hyperlink to ``cell``.

    Using a real :class:`Hyperlink` with ``location=`` (rather than the
    ``cell.hyperlink = "#'Sheet'!A1"`` string shorthand) is what Excel
    expects for "place in this document" navigation — it round-trips
    through openpyxl as ``hyperlink.location`` rather than landing on
    ``hyperlink.target``.
    """
    cell.hyperlink = Hyperlink(
        ref=cell.coordinate,
        location=f"'{target_sheet}'!{target_cell}",
        display=str(cell.value) if cell.value is not None else "",
    )


# ── v5 colour palette ────────────────────────────────────────────────
SLATE        = "2C3E50"
SLATE_LITE   = "7F8C9A"
LINK_BLUE    = "0563C1"
MUTED_RED    = "B91C1C"
MUTED_GRN    = "047857"
TOP_FILL     = "E1E7EF"
SUB_FILL     = "F2F3F5"
HAIR_GRAY    = "D6DBE0"
ALERT_BG     = "FCEAEA"
ALERT_BORDER = "F0B5B5"
OK_BG        = "E8F5EE"
OK_BORDER    = "A7D5B8"

INR_FMT = '[>=10000000]##\\,##\\,##\\,##0.00;[>=100000]##\\,##\\,##0.00;##,##0.00'

# Fieldnames the upstream report MUST expose. Looked up by name, not index.
NUM_FIELDS = (
    "opening_debit", "opening_credit",
    "debit",         "credit",
    "closing_debit", "closing_credit",
)
REQUIRED_FIELDS = ("account",) + NUM_FIELDS


# =====================================================================
# Public entry point
# =====================================================================

def build_formatted_tb(filters, columns, rows,
                        company_name=None, abbr=None):
    """Build the v5 formatted TB workbook.

    Parameters
    ----------
    filters : dict
        The TB filters from the standard report. ``company``,
        ``from_date``, ``to_date`` are read; nothing else is required.
    columns : list[dict]
        Columns as returned by ``trial_balance.execute()``. Validated
        for the presence of every fieldname in :data:`REQUIRED_FIELDS`.
    rows : list[dict]
        Rows as returned by ``trial_balance.execute()``. The last row
        is treated as the Total row.
    company_name : str, optional
        Display name for the title block. Defaults to ``filters["company"]``.
    abbr : str, optional
        Company abbreviation used to strip the suffix (e.g. ``" - JEWIPL"``)
        from every account name. If absent, no stripping happens.

    Returns
    -------
    (bytes, str)
        The xlsx bytes and the suggested filename.
    """
    if not isinstance(filters, dict):
        raise ValueError("filters must be a dict")

    _validate_columns(columns)

    company  = filters.get("company") or ""
    company_name = company_name or company
    suffix = f" - {abbr}" if abbr else ""

    from_date = _to_date(filters.get("from_date"))
    to_date   = _to_date(filters.get("to_date"))
    period    = _format_period_label(from_date, to_date)

    data, total_nums = _parse_rows(rows, suffix)

    wb = _build_workbook(company_name, period, data, total_nums)

    buf = io.BytesIO()
    wb.save(buf)
    filename = _format_filename(abbr or company, from_date, to_date)
    return buf.getvalue(), filename


# =====================================================================
# Data shaping
# =====================================================================

def _validate_columns(columns):
    """Raise if the upstream report is missing any expected fieldname."""
    field_set = {(c.get("fieldname") if isinstance(c, dict) else None)
                  for c in (columns or [])}
    missing = [f for f in REQUIRED_FIELDS if f not in field_set]
    if missing:
        raise ValueError(
            "Trial Balance report is missing expected fields: "
            + ", ".join(missing)
            + ". Has the upstream report changed?"
        )


def _parse_rows(rows, suffix):
    """Convert ERPNext's rows into the internal ``data`` list.

    Returns ``(data, total_nums)`` where ``data`` is a list of dicts
    ``{level, name, nums:[6], is_group}`` and ``total_nums`` is the
    six-element list from the last row. If ``rows`` is empty, returns
    an empty list and zero totals.
    """
    if not rows:
        return [], [0] * 6

    *body_rows, total_row = rows

    data = []
    for r in body_rows:
        if r is None:
            continue
        raw_name = r.get("account") or ""
        if not raw_name.strip():
            continue
        level = _row_indent(r)
        name  = _strip_suffix(raw_name.strip(), suffix)
        nums  = [r.get(f) for f in NUM_FIELDS]
        data.append({"level": level, "name": name, "nums": nums})

    # is_group: row is a group iff the next visible row sits at deeper indent
    for i, row in enumerate(data):
        nxt = data[i + 1] if i + 1 < len(data) else None
        row["is_group"] = bool(nxt and nxt["level"] > row["level"])

    total_nums = [total_row.get(f) for f in NUM_FIELDS]
    return data, total_nums


def _row_indent(row):
    """Per PLAN §4.5: prefer the ``indent`` field; fall back to leading
    spaces / 4."""
    indent = row.get("indent")
    if indent is not None:
        try:
            return int(indent)
        except (TypeError, ValueError):
            pass
    s = row.get("account") or ""
    m = re.match(r"^( *)", s)
    return len(m.group(1)) // 4 if m else 0


def _strip_suffix(name, suffix):
    if suffix and name.endswith(suffix):
        return name[: -len(suffix)]
    return name


# =====================================================================
# Period-label formatting
# =====================================================================

def _to_date(v):
    """Coerce a string or date-like to a :class:`datetime.date`. Returns
    ``None`` on falsy input. Accepts ISO-format strings."""
    if not v:
        return None
    if isinstance(v, _date):
        return v
    if isinstance(v, str):
        try:
            return _date.fromisoformat(v[:10])
        except ValueError:
            return None
    return None


def _format_period_label(from_date, to_date):
    """Indian fiscal year (Apr–Mar) is the only convention we render.

    - Standard FY (1-Apr to 31-Mar of following year):
      ``Trial Balance · FY YYYY-YYYY (DD Mon YYYY – DD Mon YYYY)``
    - Anything else: just the date range.
    - Missing dates: just the title.
    """
    if not (from_date and to_date):
        return "Trial Balance"

    is_indian_fy = (
        from_date.month == 4 and from_date.day == 1
        and to_date.month == 3 and to_date.day == 31
        and to_date.year == from_date.year + 1
    )

    def fmt(d):
        return d.strftime("%d %b %Y")

    if is_indian_fy:
        return (f"Trial Balance  ·  FY {from_date.year}-{to_date.year}  "
                f"({fmt(from_date)} – {fmt(to_date)})")
    return f"Trial Balance  ·  {fmt(from_date)} – {fmt(to_date)}"


def _format_filename(label, from_date, to_date):
    fd = from_date.isoformat() if from_date else ""
    td = to_date.isoformat()   if to_date   else ""
    return f"TB_{label}_{fd}_{td}.xlsx"


# =====================================================================
# Workbook builder
# =====================================================================

def _build_workbook(company_name, period_label, data, total_nums):
    wb = Workbook()
    wb.remove(wb.active)

    hair    = Side(style="hair",   color=SLATE_LITE)
    thin    = Side(style="thin",   color=SLATE_LITE)
    medium  = Side(style="medium", color=SLATE)
    double  = Side(style="double", color=SLATE)
    row_sep = Side(style="hair",   color=HAIR_GRAY)

    # Build Detail first so Summary can hyperlink to known row numbers.
    ws_d = wb.create_sheet("Detail")
    top_group_rows, total_r = _write_detail(
        ws_d, company_name, period_label, data, total_nums,
        thin=thin, medium=medium, double=double, row_sep=row_sep,
    )

    ws_s = wb.create_sheet("Summary", 0)
    _write_summary(
        ws_s, company_name, period_label,
        top_group_rows=top_group_rows,
        total_r=total_r,
        total_nums=total_nums,
        thin=thin, medium=medium, row_sep=row_sep,
    )

    wb.active = 0  # Summary opens first
    return wb


def _write_detail(ws_d, company, period, data, total_nums,
                   thin, medium, double, row_sep):
    ws_d.sheet_properties.outlinePr.summaryBelow = False
    ws_d.sheet_properties.outlinePr.summaryRight = False
    ws_d.sheet_view.showGridLines = False

    # Title block
    ws_d["A1"] = company
    ws_d["A1"].font = Font(name="Calibri", size=12, bold=True, color=SLATE)
    ws_d["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws_d.row_dimensions[1].height = 20

    ws_d["A2"] = period
    ws_d["A2"].font = Font(name="Calibri", size=10, color=SLATE_LITE)
    ws_d["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws_d.row_dimensions[2].height = 15

    back = ws_d.cell(row=2, column=7, value="←  Back to Summary")
    _intra_link(back, "Summary", "A1")
    back.font = Font(name="Calibri", size=10, color=LINK_BLUE,
                      underline="single")
    back.alignment = Alignment(horizontal="right")

    for col in range(1, 8):
        ws_d.cell(row=3, column=col).border = Border(bottom=medium)
    ws_d.row_dimensions[3].height = 4

    HDR = 4
    headers = ["Account", "Opening (Dr)", "Opening (Cr)",
                "Debit", "Credit", "Closing (Dr)", "Closing (Cr)"]
    for col, h in enumerate(headers, 1):
        c = ws_d.cell(row=HDR, column=col, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color=SLATE)
        c.alignment = Alignment(
            horizontal=("left" if col == 1 else "right"),
            vertical="center",
        )
        c.border = Border(bottom=thin)
    ws_d.row_dimensions[HDR].height = 22

    # Body rows
    top_group_rows = {}
    r = HDR + 1
    for row in data:
        is_top = (row["level"] == 0 and row["is_group"])
        if is_top:
            top_group_rows[row["name"]] = r

        a = ws_d.cell(row=r, column=1, value=row["name"])
        a.alignment = Alignment(horizontal="left", vertical="center",
                                  indent=row["level"])

        for j, v in enumerate(row["nums"], 2):
            c = ws_d.cell(row=r, column=j,
                            value=(None if v in (0, None) else v))
            c.number_format = INR_FMT
            c.alignment = Alignment(horizontal="right", vertical="center")

        if is_top:
            fill = PatternFill("solid", fgColor=TOP_FILL)
            font = Font(name="Calibri", size=11, bold=True, color=SLATE)
            for col in range(1, 8):
                cc = ws_d.cell(row=r, column=col)
                cc.font = font
                cc.fill = fill
                cc.border = Border(bottom=row_sep)
            ws_d.row_dimensions[r].height = 22

        elif row["is_group"] and row["level"] == 1:
            fill = PatternFill("solid", fgColor=SUB_FILL)
            for col in range(1, 8):
                cc = ws_d.cell(row=r, column=col)
                cc.font = Font(name="Calibri", size=10, bold=True)
                cc.fill = fill
                cc.border = Border(bottom=row_sep)
            ws_d.row_dimensions[r].height = 19

        elif row["is_group"]:
            for col in range(1, 8):
                cc = ws_d.cell(row=r, column=col)
                cc.font = Font(name="Calibri", size=10, bold=True)
                cc.border = Border(bottom=row_sep)

        else:
            for col in range(1, 8):
                cc = ws_d.cell(row=r, column=col)
                cc.font = Font(name="Calibri", size=10, color="3D3D3D")
                cc.border = Border(bottom=row_sep)

        # Outline grouping — only two levels (top=0, everything else=1)
        ws_d.row_dimensions[r].outline_level = 0 if row["level"] == 0 else 1
        r += 1

    # Total row
    total_r = r
    ws_d.cell(row=total_r, column=1, value="Total")
    for j, v in enumerate(total_nums, 2):
        c = ws_d.cell(row=total_r, column=j,
                        value=(None if v in (0, None) else v))
        c.number_format = INR_FMT
    for col in range(1, 8):
        cc = ws_d.cell(row=total_r, column=col)
        cc.font = Font(name="Calibri", size=11, bold=True, color=SLATE)
        cc.border = Border(top=medium, bottom=double)
        cc.alignment = Alignment(
            horizontal=("left" if col == 1 else "right"),
            vertical="center",
        )
    ws_d.row_dimensions[total_r].height = 22

    # Tie-check row (italic slate-lite label, italic muted-red diff)
    check_r = total_r + 2
    lbl = ws_d.cell(row=check_r, column=1,
                     value="Check  ·  Closing (Dr) − Closing (Cr)")
    lbl.font = Font(name="Calibri", size=9, italic=True, color=SLATE_LITE)
    diff = ws_d.cell(row=check_r, column=7,
                      value=f"=F{total_r}-G{total_r}")
    diff.number_format = INR_FMT
    diff.font = Font(name="Calibri", size=9, italic=True,
                      color=MUTED_RED, bold=True)
    diff.alignment = Alignment(horizontal="right")

    # Sign-off footer (per §3.2 — added per spec; absent in build_v5_reference.py)
    sign_r = check_r + 5
    sign_lbl_font = Font(name="Calibri", size=9, color=SLATE_LITE)
    for col, label in [(1, "Prepared by"), (4, "Reviewed by"), (6, "Date")]:
        c = ws_d.cell(row=sign_r, column=col, value=label)
        c.font = sign_lbl_font
        c.alignment = Alignment(horizontal="left", vertical="center")
    line_r = sign_r + 1
    for start, end in [(1, 3), (4, 5), (6, 7)]:
        for col in range(start, end + 1):
            ws_d.cell(row=line_r, column=col).border = Border(bottom=thin)
        ws_d.merge_cells(start_row=line_r, start_column=start,
                          end_row=line_r,   end_column=end)
    ws_d.row_dimensions[line_r].height = 18

    # Layout / page setup
    ws_d.column_dimensions["A"].width = 42
    for col in "BCDEFG":
        ws_d.column_dimensions[col].width = 15
    ws_d.freeze_panes = "B5"
    ws_d.print_title_rows = "1:4"
    ws_d.page_setup.orientation = ws_d.ORIENTATION_LANDSCAPE
    ws_d.page_setup.paperSize = ws_d.PAPERSIZE_A4
    ws_d.page_setup.fitToWidth = 1
    ws_d.page_setup.fitToHeight = 0
    ws_d.sheet_properties.pageSetUpPr.fitToPage = True
    ws_d.page_margins.left = 0.3
    ws_d.page_margins.right = 0.3
    ws_d.page_margins.top = 0.5
    ws_d.page_margins.bottom = 0.5

    return top_group_rows, total_r


def _write_summary(ws_s, company, period, top_group_rows, total_r,
                    total_nums, thin, medium, row_sep):
    ws_s.sheet_view.showGridLines = False

    closing_dr = float(total_nums[4] or 0)
    closing_cr = float(total_nums[5] or 0)
    tb_diff = closing_dr - closing_cr
    tb_tied = abs(tb_diff) < 0.01

    # Column widths first (merges respect them)
    ws_s.column_dimensions["A"].width = 3
    ws_s.column_dimensions["B"].width = 34
    ws_s.column_dimensions["C"].width = 2
    ws_s.column_dimensions["D"].width = 18
    ws_s.column_dimensions["E"].width = 18
    ws_s.column_dimensions["F"].width = 3

    # Title block (rows 2–4)
    ws_s.merge_cells("B2:E2")
    ws_s["B2"] = company
    ws_s["B2"].font = Font(name="Calibri", size=14, bold=True, color=SLATE)
    ws_s["B2"].alignment = Alignment(horizontal="left", vertical="center")
    ws_s.row_dimensions[2].height = 26

    ws_s.merge_cells("B3:E3")
    ws_s["B3"] = period
    ws_s["B3"].font = Font(name="Calibri", size=10, color=SLATE_LITE)
    ws_s["B3"].alignment = Alignment(horizontal="left", vertical="center")
    ws_s.row_dimensions[3].height = 16

    for col in range(2, 6):
        ws_s.cell(row=4, column=col).border = Border(bottom=medium)
    ws_s.row_dimensions[4].height = 4

    # Tie-status banner (rows 6–7) — colour computed at gen-time
    banner_bg     = OK_BG     if tb_tied else ALERT_BG
    banner_border = OK_BORDER if tb_tied else ALERT_BORDER
    banner_text   = MUTED_GRN if tb_tied else MUTED_RED
    banner_icon   = "✓" if tb_tied else "⚠"
    banner_head   = ("TRIAL BALANCE TIED" if tb_tied
                      else "TRIAL BALANCE OUT OF BALANCE")
    banner_sub    = ("Closing (Dr) equals Closing (Cr)" if tb_tied
                      else "Difference  ·  Closing (Dr) − Closing (Cr)")

    bn_side = Side(style="thin", color=banner_border)
    bn_fill = PatternFill("solid", fgColor=banner_bg)

    ws_s.merge_cells("B6:E6")
    banner = ws_s.cell(row=6, column=2,
                        value=f"  {banner_icon}   {banner_head}")
    banner.font = Font(name="Calibri", size=13, bold=True, color=banner_text)
    banner.alignment = Alignment(horizontal="left", vertical="center")

    ws_s.merge_cells("B7:D7")
    sub = ws_s.cell(row=7, column=2, value=f"    {banner_sub}")
    sub.font = Font(name="Calibri", size=10, color=SLATE)
    sub.alignment = Alignment(horizontal="left", vertical="center")

    amt = ws_s.cell(
        row=7, column=5,
        value=(0 if tb_tied else f"=Detail!F{total_r}-Detail!G{total_r}"),
    )
    amt.number_format = INR_FMT
    amt.font = Font(name="Calibri", size=12, bold=True, color=banner_text)
    amt.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    for r_ in (6, 7):
        for c_ in range(2, 6):
            cc = ws_s.cell(row=r_, column=c_)
            cc.fill = bn_fill
            top_b   = bn_side if r_ == 6 else None
            bot_b   = bn_side if r_ == 7 else None
            left_b  = bn_side if c_ == 2 else None
            right_b = bn_side if c_ == 5 else None
            cc.border = Border(top=top_b, bottom=bot_b,
                                left=left_b, right=right_b)
    ws_s.row_dimensions[6].height = 26
    ws_s.row_dimensions[7].height = 22

    # Closing Position
    sr = 10
    ws_s.cell(row=sr, column=2, value="Closing Position").font = Font(
        name="Calibri", size=11, bold=True, color=SLATE)
    ws_s.row_dimensions[sr].height = 22
    sr += 1

    hdr_fill = PatternFill("solid", fgColor=TOP_FILL)
    ws_s.cell(row=sr, column=2, value="Group")
    ws_s.cell(row=sr, column=4, value="Closing (Dr)")
    ws_s.cell(row=sr, column=5, value="Closing (Cr)")
    for c_ in range(2, 6):
        cc = ws_s.cell(row=sr, column=c_)
        cc.font = Font(name="Calibri", size=10, bold=True, color=SLATE)
        cc.fill = hdr_fill
        cc.border = Border(bottom=row_sep)
        cc.alignment = Alignment(
            horizontal=("right" if c_ in (4, 5) else "left"),
            vertical="center",
        )
    ws_s.row_dimensions[sr].height = 20
    sr += 1

    top_groups_for_summary = [
        ("Application of Funds (Assets)", "Assets"),
        ("Source of Funds (Liabilities)", "Liabilities"),
        ("Income",                         "Income"),
        ("Expenses",                       "Expenses"),
    ]
    for tg_name, display in top_groups_for_summary:
        if tg_name in top_group_rows:
            det_r = top_group_rows[tg_name]
            nm = ws_s.cell(row=sr, column=2, value=display)
            _intra_link(nm, "Detail", f"A{det_r}")
            nm.font = Font(name="Calibri", size=11, color=LINK_BLUE,
                            underline="single")
            cd = ws_s.cell(
                row=sr, column=4,
                value=f'=IF(Detail!F{det_r}=0,"—",Detail!F{det_r})',
            )
            cd.number_format = INR_FMT
            cd.font = Font(name="Calibri", size=11, color=SLATE)
            cd.alignment = Alignment(horizontal="right", vertical="center")
            cc = ws_s.cell(
                row=sr, column=5,
                value=f'=IF(Detail!G{det_r}=0,"—",Detail!G{det_r})',
            )
            cc.number_format = INR_FMT
            cc.font = Font(name="Calibri", size=11, color=SLATE)
            cc.alignment = Alignment(horizontal="right", vertical="center")
        else:
            nm = ws_s.cell(row=sr, column=2, value=display)
            nm.font = Font(name="Calibri", size=11, color=SLATE_LITE,
                            italic=True)
            for ccol in (4, 5):
                em = ws_s.cell(row=sr, column=ccol, value="—")
                em.font = Font(name="Calibri", size=11, color=SLATE_LITE)
                em.alignment = Alignment(horizontal="right", vertical="center")
        for c_ in range(2, 6):
            ws_s.cell(row=sr, column=c_).border = Border(bottom=row_sep)
        ws_s.row_dimensions[sr].height = 22
        sr += 1

    # Period Activity
    sr += 2
    ws_s.cell(row=sr, column=2, value="Period Activity").font = Font(
        name="Calibri", size=11, bold=True, color=SLATE)
    ws_s.row_dimensions[sr].height = 22
    sr += 1

    ws_s.cell(row=sr, column=2, value="Movement")
    ws_s.cell(row=sr, column=4, value="Amount")
    ws_s.merge_cells(start_row=sr, start_column=4, end_row=sr, end_column=5)
    for c_ in (2, 4, 5):
        cc = ws_s.cell(row=sr, column=c_)
        cc.font = Font(name="Calibri", size=10, bold=True, color=SLATE)
        cc.fill = hdr_fill
        cc.border = Border(bottom=row_sep)
        cc.alignment = Alignment(
            horizontal=("left" if c_ == 2 else "right"),
            vertical="center",
        )
    ws_s.row_dimensions[sr].height = 20
    sr += 1

    activity_rows = [
        ("Total Debits",   f"=Detail!D{total_r}",                     False),
        ("Total Credits",  f"=Detail!E{total_r}",                     False),
        ("Net (Dr − Cr)",  f"=Detail!D{total_r}-Detail!E{total_r}",   True),
    ]
    for lbl, formula, is_net in activity_rows:
        nm = ws_s.cell(row=sr, column=2, value=lbl)
        nm.font = Font(name="Calibri", size=11, bold=is_net,
                        color=(banner_text if is_net else "000000"))
        nm.alignment = Alignment(horizontal="left", vertical="center")

        val = ws_s.cell(row=sr, column=4, value=formula)
        val.number_format = INR_FMT
        val.font = Font(name="Calibri", size=11, bold=is_net,
                         color=(banner_text if is_net else SLATE))
        val.alignment = Alignment(horizontal="right", vertical="center")
        ws_s.merge_cells(start_row=sr, start_column=4, end_row=sr, end_column=5)

        for c_ in (2, 4, 5):
            cc_ = ws_s.cell(row=sr, column=c_)
            if is_net:
                cc_.border = Border(top=thin, bottom=row_sep)
                cc_.fill = PatternFill("solid", fgColor=banner_bg)
            else:
                cc_.border = Border(bottom=row_sep)
        ws_s.row_dimensions[sr].height = 22
        sr += 1

    # Footer link
    sr += 3
    view_d = ws_s.cell(row=sr, column=2,
                        value="View full account-level detail  →")
    _intra_link(view_d, "Detail", "A1")
    view_d.font = Font(name="Calibri", size=10, color=LINK_BLUE,
                        underline="single", italic=True)
    view_d.alignment = Alignment(horizontal="left", vertical="center")

    # Page setup
    ws_s.page_setup.orientation = ws_s.ORIENTATION_PORTRAIT
    ws_s.page_setup.paperSize = ws_s.PAPERSIZE_A4
    ws_s.page_setup.fitToWidth = 1
    ws_s.sheet_properties.pageSetUpPr.fitToPage = True
    ws_s.page_margins.left = 0.4
    ws_s.page_margins.right = 0.4
    ws_s.page_margins.top = 0.5
    ws_s.page_margins.bottom = 0.5
