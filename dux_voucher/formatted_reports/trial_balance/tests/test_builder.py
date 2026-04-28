"""
Unit tests for the formatted Trial Balance builder.

Covers the five concerns called out in PLAN.md §5.1:

1.  parsing of ERPNext-shaped rows
2.  number formatting (0 → blank)
3.  hierarchy classification (top / sub / deeper / leaf)
4.  tie computation (banner colour driver)
5.  file-output integrity (round-trip through openpyxl)

The builder is intentionally side-effect-free, so these tests do not
need a Frappe site context; they run as plain unittest cases.
"""

import io
import unittest
from datetime import date

from openpyxl import load_workbook

from dux_voucher.formatted_reports.trial_balance import builder


# =====================================================================
# Test data — synthetic columns + rows shaped like ERPNext's TB output
# =====================================================================

def _columns():
    """Mirror the standard Trial Balance report's column descriptors —
    only the fieldnames matter to the builder."""
    return [
        {"fieldname": "account",         "label": "Account"},
        {"fieldname": "opening_debit",   "label": "Opening (Dr)"},
        {"fieldname": "opening_credit",  "label": "Opening (Cr)"},
        {"fieldname": "debit",           "label": "Debit"},
        {"fieldname": "credit",          "label": "Credit"},
        {"fieldname": "closing_debit",   "label": "Closing (Dr)"},
        {"fieldname": "closing_credit",  "label": "Closing (Cr)"},
    ]


def _row(account, indent, ob_d=0, ob_c=0, dr=0, cr=0, cl_d=0, cl_c=0):
    return {
        "account":        account,
        "indent":         indent,
        "opening_debit":  ob_d,
        "opening_credit": ob_c,
        "debit":          dr,
        "credit":         cr,
        "closing_debit":  cl_d,
        "closing_credit": cl_c,
    }


def _rows_jewipl_tied():
    """Three top groups, one missing (Income), tied at 1,000,000 each."""
    return [
        _row("Application of Funds (Assets) - JEWIPL", 0,
             cl_d=1_000_000),
        _row("    Current Assets - JEWIPL",            1,
             cl_d=1_000_000),
        _row("        Cash - JEWIPL",                  2,
             dr=1_000_000, cl_d=1_000_000),
        _row("Source of Funds (Liabilities) - JEWIPL", 0,
             cl_c=600_000),
        _row("    Equity - JEWIPL",                    1,
             cl_c=600_000),
        _row("Expenses - JEWIPL", 0,
             dr=400_000, cl_d=400_000),
        # Total row (last row, by contract)
        _row("'Total'", 0,
             dr=1_400_000, cr=1_000_000,
             cl_d=1_400_000, cl_c=1_400_000),
    ]


def _rows_jewipl_out_of_balance():
    rows = _rows_jewipl_tied()
    rows[-1]["closing_credit"] = 1_399_990  # 10 short
    return rows


# =====================================================================
# Test 1 — parsing
# =====================================================================

class TestParse(unittest.TestCase):

    def test_strips_suffix(self):
        data, _ = builder._parse_rows(_rows_jewipl_tied(), suffix=" - JEWIPL")
        names = [d["name"] for d in data]
        self.assertNotIn("Cash - JEWIPL", names)
        self.assertIn("Cash", names)

    def test_indent_field_preferred_over_leading_spaces(self):
        # account string says level=2 by leading spaces, indent field says 5.
        # Builder should trust indent.
        rows = [
            _row("        Cash - JEWIPL",  5),
            _row("'Total'",                 0),
        ]
        data, _ = builder._parse_rows(rows, suffix=" - JEWIPL")
        self.assertEqual(data[0]["level"], 5)

    def test_falls_back_to_leading_spaces_when_indent_absent(self):
        rows = [
            {"account": "        Cash"},  # 8 spaces → level 2
            {"account": "'Total'"},
        ]
        data, _ = builder._parse_rows(rows, suffix="")
        self.assertEqual(data[0]["level"], 2)

    def test_total_nums_come_from_last_row(self):
        _, totals = builder._parse_rows(_rows_jewipl_tied(),
                                          suffix=" - JEWIPL")
        # NUM_FIELDS order: opening_debit, opening_credit, debit, credit,
        # closing_debit, closing_credit
        self.assertEqual(totals[2], 1_400_000)  # debit
        self.assertEqual(totals[4], 1_400_000)  # closing_debit

    def test_empty_rows_yield_empty_data(self):
        data, totals = builder._parse_rows([], suffix=" - JEWIPL")
        self.assertEqual(data, [])
        self.assertEqual(totals, [0] * 6)

    def test_skips_blank_account_rows(self):
        rows = [
            _row("Application of Funds (Assets) - JEWIPL", 0),
            {"account": "", "indent": 0},
            _row("    Current Assets - JEWIPL", 1),
            _row("'Total'", 0),
        ]
        data, _ = builder._parse_rows(rows, suffix=" - JEWIPL")
        # Two body rows survive; blank one is dropped.
        self.assertEqual(len(data), 2)


# =====================================================================
# Test 2 — number formatting (0 → None in cell value)
# =====================================================================

class TestNumberFormatting(unittest.TestCase):

    def test_zero_becomes_blank_in_workbook(self):
        xlsx_bytes, _ = builder.build_formatted_tb(
            filters={"company": "Acme",
                      "from_date": "2026-04-01",
                      "to_date":   "2027-03-31"},
            columns=_columns(),
            rows=_rows_jewipl_tied(),
            company_name="Acme Pvt Ltd",
            abbr="JEWIPL",
        )
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws_d = wb["Detail"]
        # Row 5 = first data row (Application of Funds (Assets)) → Opening
        # Debit (col B) is zero per fixture, so should be None in the cell.
        self.assertIsNone(ws_d.cell(row=5, column=2).value)
        # Closing debit (col F) is 1_000_000 → numeric value preserved.
        self.assertEqual(ws_d.cell(row=5, column=6).value, 1_000_000)

    def test_indian_number_format_applied_to_amount_cells(self):
        xlsx_bytes, _ = builder.build_formatted_tb(
            filters={"company": "Acme",
                      "from_date": "2026-04-01",
                      "to_date":   "2027-03-31"},
            columns=_columns(),
            rows=_rows_jewipl_tied(),
            company_name="Acme Pvt Ltd",
            abbr="JEWIPL",
        )
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws_d = wb["Detail"]
        self.assertEqual(ws_d.cell(row=5, column=6).number_format,
                          builder.INR_FMT)


# =====================================================================
# Test 3 — hierarchy classification
# =====================================================================

class TestHierarchy(unittest.TestCase):

    def test_top_group_detected(self):
        data, _ = builder._parse_rows(_rows_jewipl_tied(),
                                        suffix=" - JEWIPL")
        # "Application of Funds (Assets)" sits at level 0 with deeper children
        first = data[0]
        self.assertEqual(first["name"], "Application of Funds (Assets)")
        self.assertEqual(first["level"], 0)
        self.assertTrue(first["is_group"])

    def test_leaf_detected(self):
        data, _ = builder._parse_rows(_rows_jewipl_tied(),
                                        suffix=" - JEWIPL")
        cash = next(d for d in data if d["name"] == "Cash")
        self.assertEqual(cash["level"], 2)
        # Cash is followed by another level-0 group — same-or-shallower → not a group
        self.assertFalse(cash["is_group"])

    def test_top_group_with_no_children_is_not_group(self):
        # "Expenses" in our fixture has no children — it sits at level 0 but
        # is followed by the Total row. is_group should be False.
        data, _ = builder._parse_rows(_rows_jewipl_tied(),
                                        suffix=" - JEWIPL")
        expenses = next(d for d in data if d["name"] == "Expenses")
        self.assertFalse(expenses["is_group"])


# =====================================================================
# Test 4 — tie computation (banner colour driver)
# =====================================================================

class TestTieComputation(unittest.TestCase):

    def _summary_banner(self, rows):
        xlsx_bytes, _ = builder.build_formatted_tb(
            filters={"company": "Acme",
                      "from_date": "2026-04-01",
                      "to_date":   "2027-03-31"},
            columns=_columns(),
            rows=rows,
            company_name="Acme Pvt Ltd",
            abbr="JEWIPL",
        )
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws_s = wb["Summary"]
        return ws_s["B6"].value

    def test_tied_shows_check_mark(self):
        banner = self._summary_banner(_rows_jewipl_tied())
        self.assertIn("✓",                     banner)
        self.assertIn("TRIAL BALANCE TIED",    banner)

    def test_out_of_balance_shows_warning(self):
        banner = self._summary_banner(_rows_jewipl_out_of_balance())
        self.assertIn("⚠",                              banner)
        self.assertIn("TRIAL BALANCE OUT OF BALANCE",   banner)


# =====================================================================
# Test 5 — file-output integrity
# =====================================================================

class TestFileIntegrity(unittest.TestCase):

    def setUp(self):
        self.xlsx_bytes, self.filename = builder.build_formatted_tb(
            filters={"company":  "Acme",
                      "from_date": "2026-04-01",
                      "to_date":   "2027-03-31"},
            columns=_columns(),
            rows=_rows_jewipl_tied(),
            company_name="Acme Pvt Ltd",
            abbr="JEWIPL",
        )
        self.wb = load_workbook(io.BytesIO(self.xlsx_bytes))

    def test_has_summary_and_detail_sheets(self):
        self.assertIn("Summary", self.wb.sheetnames)
        self.assertIn("Detail",  self.wb.sheetnames)

    def test_summary_opens_first(self):
        # Active sheet should be Summary
        self.assertEqual(self.wb.active.title, "Summary")

    def test_filename_pattern(self):
        self.assertEqual(self.filename,
                          "TB_JEWIPL_2026-04-01_2027-03-31.xlsx")

    def test_summary_cross_references_detail_total_row(self):
        # The amount cell in the banner (E7) — when out of balance — should
        # carry the formula =Detail!F<total>-Detail!G<total>. Tied case is
        # tested separately as it short-circuits to literal 0.
        xlsx_bytes, _ = builder.build_formatted_tb(
            filters={"company":  "Acme",
                      "from_date": "2026-04-01",
                      "to_date":   "2027-03-31"},
            columns=_columns(),
            rows=_rows_jewipl_out_of_balance(),
            company_name="Acme Pvt Ltd",
            abbr="JEWIPL",
        )
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws_s = wb["Summary"]
        formula = ws_s["E7"].value
        self.assertTrue(isinstance(formula, str)
                         and formula.startswith("=Detail!F")
                         and "-Detail!G" in formula,
                         f"unexpected banner formula: {formula!r}")

    def test_top_group_rows_are_hyperlinked_from_summary(self):
        ws_s = self.wb["Summary"]
        # Closing Position table starts at row 11 (header) + 1 = 12 for first
        # data row (Assets). The cell value is the *display* string; hyperlink
        # is on the cell.hyperlink attribute.
        # Find the row with value "Assets" in column B (sweep rows 12–16).
        found = None
        for r in range(12, 18):
            if ws_s.cell(row=r, column=2).value == "Assets":
                found = ws_s.cell(row=r, column=2)
                break
        self.assertIsNotNone(found, "Assets row not found in Summary")
        self.assertIsNotNone(found.hyperlink,
                              "Assets row should hyperlink to Detail")
        self.assertIn("Detail", str(found.hyperlink.location or ""))

    def test_missing_top_group_renders_em_dash_without_hyperlink(self):
        # The fixture has no Income top-group → Summary should show "Income"
        # in italic with em-dashes, NO hyperlink.
        ws_s = self.wb["Summary"]
        for r in range(12, 18):
            if ws_s.cell(row=r, column=2).value == "Income":
                income_cell = ws_s.cell(row=r, column=2)
                self.assertIsNone(income_cell.hyperlink)
                self.assertEqual(ws_s.cell(row=r, column=4).value, "—")
                self.assertEqual(ws_s.cell(row=r, column=5).value, "—")
                return
        self.fail("Income row not found in Summary")

    def test_detail_total_row_uses_strong_borders(self):
        ws_d = self.wb["Detail"]
        # Walk down column A to find the "Total" row
        for r in range(5, 50):
            if ws_d.cell(row=r, column=1).value == "Total":
                # Top border medium, bottom border double — per spec
                top = ws_d.cell(row=r, column=2).border.top
                bot = ws_d.cell(row=r, column=2).border.bottom
                self.assertEqual(top.style, "medium")
                self.assertEqual(bot.style, "double")
                return
        self.fail("Total row not found in Detail sheet")


# =====================================================================
# Period-label edge cases
# =====================================================================

class TestPeriodLabel(unittest.TestCase):

    def test_indian_fy_format(self):
        out = builder._format_period_label(
            date(2026, 4, 1), date(2027, 3, 31))
        self.assertIn("FY 2026-2027", out)
        self.assertIn("01 Apr 2026", out)
        self.assertIn("31 Mar 2027", out)

    def test_non_fy_falls_back_to_date_range(self):
        out = builder._format_period_label(
            date(2026, 1, 1), date(2026, 6, 30))
        self.assertNotIn("FY",          out)
        self.assertIn("01 Jan 2026",    out)
        self.assertIn("30 Jun 2026",    out)

    def test_missing_dates(self):
        self.assertEqual(builder._format_period_label(None, None),
                          "Trial Balance")

    def test_to_date_accepts_iso_string(self):
        self.assertEqual(builder._to_date("2026-04-01"),
                          date(2026, 4, 1))

    def test_to_date_accepts_datetime_string(self):
        # ERPNext sometimes hands dates as 'YYYY-MM-DD HH:MM:SS' strings
        self.assertEqual(builder._to_date("2026-04-01 00:00:00"),
                          date(2026, 4, 1))


# =====================================================================
# Column validation
# =====================================================================

class TestColumnValidation(unittest.TestCase):

    def test_missing_field_raises_with_message(self):
        bad_columns = [c for c in _columns()
                        if c["fieldname"] != "closing_credit"]
        with self.assertRaises(ValueError) as ctx:
            builder.build_formatted_tb(
                filters={"company": "Acme"},
                columns=bad_columns,
                rows=_rows_jewipl_tied(),
                abbr="JEWIPL",
            )
        self.assertIn("closing_credit", str(ctx.exception))


if __name__ == "__main__":
    unittest.main()
